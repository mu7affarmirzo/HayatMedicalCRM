from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, F
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from HayatMedicalCRM.auth.decorators import warehouse_manager_required
from core.models import (
    MedicationExpenseModel,
    MedicationsInStockModel,
    MedicationModel,
    Warehouse,
    AccountTransferModel,
    AccountTransferItemsModel
)
from ..forms.expense_form import MedicationExpenseForm
from itertools import chain
from operator import attrgetter


@warehouse_manager_required
def expense_list(request):
    """
    View to display list of medication expenses and account transfers (staff transfers)
    """
    # Filter parameters
    expense_type = request.GET.get('expense_type')
    warehouse_id = request.GET.get('warehouse')
    medication_name = request.GET.get('medication')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    show_type = request.GET.get('show_type', 'all')  # all, expenses, transfers

    # Fetch medication expenses
    expenses = MedicationExpenseModel.objects.select_related(
        'stock_item',
        'stock_item__item',
        'stock_item__warehouse',
        'created_by'
    )

    # Apply filters to expenses
    if expense_type and show_type != 'transfers':
        expenses = expenses.filter(expense_type=expense_type)
    if warehouse_id:
        expenses = expenses.filter(stock_item__warehouse_id=warehouse_id)
    if medication_name:
        expenses = expenses.filter(stock_item__item__name__icontains=medication_name)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)

    # Fetch account transfers (only accepted ones)
    account_transfers = AccountTransferModel.objects.filter(
        state='принято'
    ).select_related('sender', 'receiver', 'created_by')

    # Apply filters to account transfers
    if warehouse_id:
        account_transfers = account_transfers.filter(sender_id=warehouse_id)
    if date_from:
        account_transfers = account_transfers.filter(created_at__date__gte=date_from)
    if date_to:
        account_transfers = account_transfers.filter(created_at__date__lte=date_to)

    # Create unified list
    if show_type == 'expenses':
        combined_list = list(expenses)
    elif show_type == 'transfers':
        combined_list = list(account_transfers)
    else:  # all
        # Add type annotation to distinguish in template
        for exp in expenses:
            exp.item_type = 'expense'
        for transfer in account_transfers:
            transfer.item_type = 'transfer'

        combined_list = sorted(
            chain(expenses, account_transfers),
            key=lambda x: x.expense_date if hasattr(x, 'expense_date') else x.created_at,
            reverse=True
        )

    # Pagination
    paginator = Paginator(combined_list, 20)  # Show 20 items per page
    page = request.GET.get('page')
    items_page = paginator.get_page(page)

    # Calculate statistics
    total_expenses = expenses.count()
    total_transfers = account_transfers.count()
    total_combined = total_expenses + total_transfers

    # Get expense type statistics
    expense_type_stats = expenses.values('expense_type').annotate(
        count=Sum('quantity')
    )

    # Get context data
    warehouses = Warehouse.objects.all()
    expense_type_choices = MedicationExpenseModel.EXPENSE_TYPE_CHOICES

    context = {
        'items': items_page,
        'warehouses': warehouses,
        'expense_type_choices': expense_type_choices,
        'total_expenses': total_expenses,
        'total_transfers': total_transfers,
        'total_combined': total_combined,
        'expense_type_stats': expense_type_stats,
        'selected_warehouse': warehouse_id,
        'selected_expense_type': expense_type,
        'selected_show_type': show_type,
    }

    return render(request, 'warehouse/expenses/expense_list.html', context)


@warehouse_manager_required
def expense_create(request):
    """
    View to create a new medication expense and update stock accordingly
    """
    if request.method == 'POST':
        form = MedicationExpenseForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create the expense record
                    expense = form.save(commit=False)
                    expense.created_by = request.user
                    expense.save()

                    # Update the stock item
                    stock_item = expense.stock_item
                    in_pack = stock_item.item.in_pack

                    # Calculate total units to deduct
                    units_to_deduct = (expense.quantity * in_pack) + expense.unit_quantity

                    # Calculate current total units in stock
                    current_total_units = (stock_item.quantity * in_pack) + stock_item.unit_quantity

                    # Deduct from stock
                    new_total_units = current_total_units - units_to_deduct

                    # Update stock quantities
                    stock_item.quantity = new_total_units // in_pack
                    stock_item.unit_quantity = new_total_units % in_pack
                    stock_item.save()

                    messages.success(
                        request,
                        f'Расход успешно создан. Списано: {expense.quantity} уп. {expense.unit_quantity} ед. '
                        f'медикамента "{stock_item.item.name}". '
                        f'Осталось на складе: {stock_item.quantity} уп. {stock_item.unit_quantity} ед.'
                    )
                    return redirect('expense_list')

            except Exception as e:
                messages.error(request, f'Ошибка при создании расхода: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = MedicationExpenseForm()

    # Get warehouses for the filter dropdown
    warehouses = Warehouse.objects.all()

    context = {
        'form': form,
        'warehouses': warehouses,
    }

    return render(request, 'warehouse/expenses/expense_create.html', context)


@warehouse_manager_required
def expense_detail(request, pk):
    """
    View to display details of a specific expense
    """
    expense = get_object_or_404(
        MedicationExpenseModel.objects.select_related(
            'stock_item',
            'stock_item__item',
            'stock_item__warehouse',
            'created_by'
        ),
        pk=pk
    )

    context = {
        'expense': expense,
    }

    return render(request, 'warehouse/expenses/expense_detail.html', context)


@warehouse_manager_required
def expense_export_excel(request):
    """
    Export expenses to Excel file
    """
    # Get filtered expenses
    expenses = MedicationExpenseModel.objects.select_related(
        'stock_item',
        'stock_item__item',
        'stock_item__warehouse',
        'created_by'
    ).order_by('-expense_date')

    # Apply same filters as list view
    expense_type = request.GET.get('expense_type')
    warehouse_id = request.GET.get('warehouse')
    medication_name = request.GET.get('medication')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if expense_type:
        expenses = expenses.filter(expense_type=expense_type)
    if warehouse_id:
        expenses = expenses.filter(stock_item__warehouse_id=warehouse_id)
    if medication_name:
        expenses = expenses.filter(stock_item__item__name__icontains=medication_name)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы медикаментов"

    # Define headers
    headers = [
        'Дата расхода',
        'Медикамент',
        'Склад',
        'Тип расхода',
        'Количество упаковок',
        'Количество единиц',
        'Всего единиц',
        'Стоимость',
        'Серия прихода',
        'Срок годности',
        'Создано',
        'Пользователь',
        'Примечание'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, expense in enumerate(expenses, 2):
        ws.cell(row=row_num, column=1, value=expense.expense_date.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=2, value=expense.stock_item.item.name)
        ws.cell(row=row_num, column=3, value=expense.stock_item.warehouse.name)
        ws.cell(row=row_num, column=4, value=expense.get_expense_type_display())
        ws.cell(row=row_num, column=5, value=expense.quantity)
        ws.cell(row=row_num, column=6, value=expense.unit_quantity)
        ws.cell(row=row_num, column=7, value=expense.total_units)
        ws.cell(row=row_num, column=8, value=expense.total_value)
        ws.cell(row=row_num, column=9, value=expense.stock_item.income_seria or '')
        ws.cell(row=row_num, column=10, value=expense.stock_item.expire_date.strftime('%d.%m.%Y') if expense.stock_item.expire_date else '')
        ws.cell(row=row_num, column=11, value=expense.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row_num, column=12, value=str(expense.created_by))
        ws.cell(row=row_num, column=13, value=expense.reason)

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=medication_expenses.xlsx'
    wb.save(response)

    return response
